from __future__ import annotations

from collections import deque
from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
import csv
import json
import re
import time
from typing import Any
from urllib import error as url_error
from urllib import parse as url_parse
from urllib import request as url_request


SUPPORTED_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".csv",
    ".json",
    ".jsonl",
    ".xlsx",
    ".pdf",
}

NOTION_DEFAULT_BASE_URL = "https://api.notion.com"
NOTION_DEFAULT_VERSION = "2022-06-28"


@dataclass(slots=True)
class LegacyImportResult:
    records: list[dict[str, Any]]
    skipped_files: list[dict[str, str]]
    parser_warnings: list[str]


@dataclass(slots=True)
class LegacySeedPlanResult:
    records: list[dict[str, Any]]
    summary: dict[str, Any]


class LegacySeedOrchestrator:
    """Build deterministic wiki-seed hints + provenance for legacy-import records."""

    def __init__(self, *, default_page_prefix: str = "legacy", max_summary_pages: int = 25) -> None:
        self.default_page_prefix = _slugify_key(default_page_prefix) or "legacy"
        self.max_summary_pages = max(1, int(max_summary_pages))

    def apply(
        self,
        *,
        records: list[dict[str, Any]],
        source_type: str,
        source_ref: str,
        project_id: str | None = None,
        source_id: str | None = None,
        run_id: str | None = None,
        config: dict[str, Any] | None = None,
    ) -> LegacySeedPlanResult:
        settings = dict(config or {})
        page_prefix = _slugify_key(str(settings.get("seed_page_prefix") or self.default_page_prefix))
        if not page_prefix:
            page_prefix = self.default_page_prefix
        space_override = _slugify_key(str(settings.get("seed_space_key") or ""))
        group_mode = str(settings.get("seed_group_mode") or "category_entity").strip().lower()
        if group_mode not in {"entity", "category", "category_entity"}:
            group_mode = "category_entity"
        section_overrides = _normalize_section_overrides(settings.get("seed_section_overrides"))

        record_out: list[dict[str, Any]] = []
        page_stats: dict[str, dict[str, Any]] = {}
        source_ref_fingerprint = sha256(str(source_ref).encode("utf-8")).hexdigest()[:16]

        for index, raw_record in enumerate(records, start=1):
            item = dict(raw_record)
            content = _normalize_whitespace(str(item.get("content") or ""))
            if not content:
                continue
            metadata = dict(item.get("metadata") or {})

            category = _normalize_seed_category(str(item.get("category") or metadata.get("category") or ""))
            if not category:
                category = _infer_category(str(metadata.get("file_path") or item.get("source_id") or source_ref or "legacy"))

            entity_candidate = (
                str(item.get("entity_key") or metadata.get("entity_key") or "").strip()
                or _infer_seed_entity(metadata=metadata, source_id=str(item.get("source_id") or ""))
            )
            entity_key = _slugify_key(entity_candidate) or f"legacy_entity_{index}"
            space_key = space_override or _infer_seed_space(source_type=source_type, category=category, metadata=metadata)
            section_key, section_heading = _resolve_seed_section(category=category, overrides=section_overrides)
            page_token = _build_seed_page_token(group_mode=group_mode, entity_key=entity_key, category=category, metadata=metadata)
            page_slug = _join_slug(page_prefix, space_key, page_token)
            page_title = _build_seed_page_title(
                entity_key=entity_key,
                category=category,
                metadata=metadata,
                section_heading=section_heading,
            )

            content_fingerprint = str(metadata.get("sync_content_fingerprint") or "").strip() or sha256(content.encode("utf-8")).hexdigest()
            source_record_id = str(item.get("source_id") or f"legacy:{index}:{content_fingerprint[:10]}")
            item["source_id"] = source_record_id
            item["entity_key"] = entity_key
            item["category"] = category

            metadata["entity_key"] = entity_key
            metadata["category"] = category
            metadata["synapse_seed_plan"] = {
                "version": "v1",
                "space_key": space_key,
                "page_slug": page_slug,
                "page_title": page_title,
                "section_key": section_key,
                "section_heading": section_heading,
                "group_mode": group_mode,
            }
            metadata["synapse_source_provenance"] = {
                "version": "v1",
                "source_type": source_type,
                "source_ref": source_ref,
                "source_ref_fingerprint": source_ref_fingerprint,
                "project_id": project_id,
                "source_id": source_id,
                "run_id": run_id,
                "record_source_id": source_record_id,
                "record_index": index,
                "content_fingerprint": content_fingerprint[:16],
            }
            item["metadata"] = metadata
            record_out.append(item)

            stat = page_stats.get(page_slug)
            if stat is None:
                stat = {
                    "page_slug": page_slug,
                    "page_title": page_title,
                    "space_key": space_key,
                    "records": 0,
                    "categories": set(),
                    "section_keys": set(),
                }
                page_stats[page_slug] = stat
            stat["records"] += 1
            stat["categories"].add(category)
            stat["section_keys"].add(section_key)

        top_pages = sorted(page_stats.values(), key=lambda item: (-int(item["records"]), str(item["page_slug"])))
        top_pages = top_pages[: self.max_summary_pages]
        summary = {
            "seed_records": len(record_out),
            "seed_pages": len(page_stats),
            "group_mode": group_mode,
            "page_prefix": page_prefix,
            "top_pages": [
                {
                    "page_slug": str(item["page_slug"]),
                    "page_title": str(item["page_title"]),
                    "space_key": str(item["space_key"]),
                    "records": int(item["records"]),
                    "categories": sorted(str(x) for x in item["categories"]),
                    "section_keys": sorted(str(x) for x in item["section_keys"]),
                }
                for item in top_pages
            ],
        }
        return LegacySeedPlanResult(records=record_out, summary=summary)


class NotionApiError(RuntimeError):
    pass


@dataclass(slots=True)
class NotionImportConfig:
    token: str
    base_url: str = NOTION_DEFAULT_BASE_URL
    notion_version: str = NOTION_DEFAULT_VERSION
    timeout_sec: int = 20
    max_retries: int = 3


class LegacyImporter:
    def __init__(
        self,
        *,
        root_dir: Path,
        max_chars_per_record: int = 1800,
        max_records_per_file: int = 500,
        include_extensions: set[str] | None = None,
    ) -> None:
        self.root_dir = root_dir
        self.max_chars_per_record = max(200, int(max_chars_per_record))
        self.max_records_per_file = max(1, int(max_records_per_file))
        self.include_extensions = include_extensions or set(SUPPORTED_EXTENSIONS)
        self._warnings: list[str] = []

    def collect_records(self, *, max_records: int | None = None) -> LegacyImportResult:
        if not self.root_dir.exists():
            raise FileNotFoundError(f"legacy import path does not exist: {self.root_dir}")

        records: list[dict[str, Any]] = []
        skipped_files: list[dict[str, str]] = []
        seen_fingerprints: set[str] = set()

        files = [p for p in sorted(self.root_dir.rglob("*")) if p.is_file()]
        for file_path in files:
            ext = file_path.suffix.lower()
            if ext not in self.include_extensions:
                continue
            try:
                chunks = self._parse_file(file_path, ext)
            except Exception as exc:
                skipped_files.append(
                    {
                        "path": str(file_path),
                        "reason": f"parse_error:{type(exc).__name__}",
                    }
                )
                continue

            rel_path = str(file_path.relative_to(self.root_dir))
            category_hint = _infer_category(rel_path)
            entity_hint = _infer_entity(rel_path)
            limited_chunks = chunks[: self.max_records_per_file]
            for idx, chunk in enumerate(limited_chunks, start=1):
                content = _normalize_whitespace(chunk)
                if not content:
                    continue
                if len(content) > self.max_chars_per_record:
                    content = f"{content[:self.max_chars_per_record]}...(truncated)"
                fingerprint = sha256(content.encode("utf-8")).hexdigest()
                if fingerprint in seen_fingerprints:
                    continue
                seen_fingerprints.add(fingerprint)
                source_id = f"{rel_path}:{idx}:{fingerprint[:10]}"
                records.append(
                    {
                        "source_id": source_id,
                        "content": content,
                        "entity_key": entity_hint,
                        "category": category_hint,
                        "metadata": {
                            "legacy_import": True,
                            "file_path": rel_path,
                            "file_ext": ext,
                            "chunk_index": idx,
                            "fingerprint": fingerprint[:16],
                        },
                    }
                )
                if max_records is not None and len(records) >= max_records:
                    return LegacyImportResult(records=records, skipped_files=skipped_files, parser_warnings=list(self._warnings))
        return LegacyImportResult(records=records, skipped_files=skipped_files, parser_warnings=list(self._warnings))

    def _parse_file(self, path: Path, ext: str) -> list[str]:
        if ext in {".txt", ".md", ".markdown"}:
            return _split_paragraphs(path.read_text(encoding="utf-8", errors="ignore"))
        if ext == ".csv":
            return self._parse_csv(path)
        if ext in {".json", ".jsonl"}:
            return self._parse_json(path, jsonl=(ext == ".jsonl"))
        if ext == ".xlsx":
            return self._parse_xlsx(path)
        if ext == ".pdf":
            return self._parse_pdf(path)
        return []

    def _parse_csv(self, path: Path) -> list[str]:
        rows: list[str] = []
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as fh:
            reader = csv.DictReader(fh)
            if reader.fieldnames:
                for row in reader:
                    serialized = "; ".join(
                        f"{key}={_normalize_whitespace(str(value or ''))}"
                        for key, value in row.items()
                        if key is not None and str(value or "").strip()
                    )
                    if serialized:
                        rows.append(serialized)
            else:
                fh.seek(0)
                raw_reader = csv.reader(fh)
                for row in raw_reader:
                    serialized = "; ".join(_normalize_whitespace(cell) for cell in row if _normalize_whitespace(cell))
                    if serialized:
                        rows.append(serialized)
        return rows

    def _parse_json(self, path: Path, *, jsonl: bool) -> list[str]:
        if jsonl:
            chunks: list[str] = []
            for line_no, line in enumerate(path.read_text(encoding="utf-8", errors="ignore").splitlines(), start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    self._warnings.append(f"{path}: invalid JSONL at line {line_no}")
                    continue
                chunks.extend(_flatten_payload(payload))
            return chunks
        payload = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
        return _flatten_payload(payload)

    def _parse_xlsx(self, path: Path) -> list[str]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            self._warnings.append(f"{path}: openpyxl not installed, skipping .xlsx parsing")
            return []
        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks: list[str] = []
        for sheet in workbook.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if not cells:
                    continue
                chunks.append(f"{sheet.title} row {row_idx}: " + " | ".join(cells))
        return chunks

    def _parse_pdf(self, path: Path) -> list[str]:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception:
            self._warnings.append(f"{path}: pypdf not installed, skipping .pdf parsing")
            return []
        reader = PdfReader(str(path))
        chunks: list[str] = []
        for idx, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            normalized = _normalize_whitespace(text)
            if normalized:
                chunks.append(f"page {idx}: {normalized}")
        return chunks


class NotionApiClient:
    def __init__(self, config: NotionImportConfig) -> None:
        self.config = config

    def get_page(self, page_id: str) -> dict[str, Any]:
        page_id = _normalize_notion_id(page_id)
        return self._request_json("GET", f"/v1/pages/{page_id}")

    def get_block_children(self, block_id: str) -> list[dict[str, Any]]:
        block_id = _normalize_notion_id(block_id)
        return self._iterate_paginated("GET", f"/v1/blocks/{block_id}/children")

    def query_database(self, database_id: str, *, max_items: int | None = None) -> list[dict[str, Any]]:
        database_id = _normalize_notion_id(database_id)
        return self._iterate_paginated(
            "POST",
            f"/v1/databases/{database_id}/query",
            payload={},
            max_items=max_items,
        )

    def _iterate_paginated(
        self,
        method: str,
        path: str,
        *,
        payload: dict[str, Any] | None = None,
        max_items: int | None = None,
    ) -> list[dict[str, Any]]:
        start_cursor: str | None = None
        out: list[dict[str, Any]] = []
        while True:
            if method == "GET":
                query: dict[str, str] = {"page_size": "100"}
                if start_cursor:
                    query["start_cursor"] = start_cursor
                result = self._request_json(method, path, query=query)
            else:
                page_payload: dict[str, Any] = dict(payload or {})
                page_payload["page_size"] = 100
                if start_cursor:
                    page_payload["start_cursor"] = start_cursor
                result = self._request_json(method, path, payload=page_payload)

            page_items = result.get("results") or []
            if isinstance(page_items, list):
                out.extend(item for item in page_items if isinstance(item, dict))
            if max_items is not None and len(out) >= max_items:
                return out[:max_items]

            if not bool(result.get("has_more")):
                break
            next_cursor = result.get("next_cursor")
            if not isinstance(next_cursor, str) or not next_cursor:
                break
            start_cursor = next_cursor
        return out

    def _request_json(
        self,
        method: str,
        path: str,
        *,
        payload: dict[str, Any] | None = None,
        query: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        url = self.config.base_url.rstrip("/") + path
        if query:
            url = f"{url}?{url_parse.urlencode(query)}"
        headers = {
            "Authorization": f"Bearer {self.config.token}",
            "Notion-Version": self.config.notion_version,
            "Content-Type": "application/json",
            "User-Agent": "synapse-legacy-import/0.1",
        }
        body = json.dumps(payload).encode("utf-8") if payload is not None else None

        for attempt in range(self.config.max_retries + 1):
            req = url_request.Request(url=url, data=body, method=method, headers=headers)
            try:
                with url_request.urlopen(req, timeout=self.config.timeout_sec) as resp:
                    raw = resp.read().decode("utf-8")
                data = json.loads(raw) if raw else {}
                if not isinstance(data, dict):
                    raise NotionApiError("notion response is not a JSON object")
                return data
            except url_error.HTTPError as exc:
                response_body = exc.read().decode("utf-8", errors="ignore")
                parsed_body: dict[str, Any] = {}
                try:
                    maybe_parsed = json.loads(response_body)
                    if isinstance(maybe_parsed, dict):
                        parsed_body = maybe_parsed
                except json.JSONDecodeError:
                    parsed_body = {}
                if exc.code == 429 and attempt < self.config.max_retries:
                    retry_after_raw = exc.headers.get("Retry-After") if exc.headers else None
                    retry_after = 1.0
                    if retry_after_raw:
                        try:
                            retry_after = max(0.2, float(retry_after_raw))
                        except ValueError:
                            retry_after = 1.0
                    time.sleep(retry_after)
                    continue
                if 500 <= exc.code <= 599 and attempt < self.config.max_retries:
                    time.sleep(0.4 * (2**attempt))
                    continue
                message = parsed_body.get("message") or response_body or exc.reason
                raise NotionApiError(f"notion api error {exc.code}: {message}") from exc
            except (url_error.URLError, TimeoutError) as exc:
                if attempt < self.config.max_retries:
                    time.sleep(0.4 * (2**attempt))
                    continue
                raise NotionApiError(f"notion request failed: {exc}") from exc


class NotionImporter:
    def __init__(
        self,
        *,
        client: NotionApiClient,
        max_chars_per_record: int = 1800,
        max_records_per_page: int = 200,
        max_pages: int = 500,
    ) -> None:
        self.client = client
        self.max_chars_per_record = max(200, int(max_chars_per_record))
        self.max_records_per_page = max(1, int(max_records_per_page))
        self.max_pages = max(1, int(max_pages))
        self._warnings: list[str] = []

    def collect_from_root_page(self, root_page_id: str, *, max_records: int | None = None) -> LegacyImportResult:
        records: list[dict[str, Any]] = []
        skipped_files: list[dict[str, str]] = []
        seen_fingerprints: set[str] = set()
        visited_pages: set[str] = set()
        queue: deque[str] = deque([_normalize_notion_id(root_page_id)])

        while queue and len(visited_pages) < self.max_pages:
            page_id = queue.popleft()
            if page_id in visited_pages:
                continue
            visited_pages.add(page_id)
            try:
                page = self.client.get_page(page_id)
            except Exception as exc:
                skipped_files.append(
                    {
                        "path": f"notion:page:{page_id}",
                        "reason": f"fetch_error:{type(exc).__name__}",
                    }
                )
                continue

            page_records, child_pages = self._records_from_page(
                page=page,
                seen_fingerprints=seen_fingerprints,
            )
            records.extend(page_records)
            for child_page_id in child_pages:
                if child_page_id not in visited_pages and child_page_id not in queue and len(visited_pages) + len(queue) < self.max_pages:
                    queue.append(child_page_id)

            if max_records is not None and len(records) >= max_records:
                return LegacyImportResult(
                    records=records[:max_records],
                    skipped_files=skipped_files,
                    parser_warnings=list(self._warnings),
                )

        if queue:
            self._warnings.append(f"notion traversal reached max_pages={self.max_pages}, remaining pages skipped")

        return LegacyImportResult(
            records=records,
            skipped_files=skipped_files,
            parser_warnings=list(self._warnings),
        )

    def collect_from_database(self, database_id: str, *, max_records: int | None = None) -> LegacyImportResult:
        records: list[dict[str, Any]] = []
        skipped_files: list[dict[str, str]] = []
        seen_fingerprints: set[str] = set()
        try:
            pages = self.client.query_database(_normalize_notion_id(database_id), max_items=self.max_pages)
        except Exception as exc:
            raise NotionApiError(f"failed to query notion database {database_id}: {exc}") from exc

        for page in pages:
            page_id = _normalize_notion_id(str(page.get("id", "")))
            if not page_id:
                skipped_files.append({"path": "notion:page:unknown", "reason": "missing_page_id"})
                continue
            page_records, _ = self._records_from_page(page=page, seen_fingerprints=seen_fingerprints)
            records.extend(page_records)
            if max_records is not None and len(records) >= max_records:
                return LegacyImportResult(
                    records=records[:max_records],
                    skipped_files=skipped_files,
                    parser_warnings=list(self._warnings),
                )

        return LegacyImportResult(
            records=records,
            skipped_files=skipped_files,
            parser_warnings=list(self._warnings),
        )

    def _records_from_page(
        self,
        *,
        page: dict[str, Any],
        seen_fingerprints: set[str],
    ) -> tuple[list[dict[str, Any]], list[str]]:
        page_id = _normalize_notion_id(str(page.get("id", "")))
        if not page_id:
            return [], []

        title = _notion_page_title(page) or f"page_{page_id[:8]}"
        page_url = str(page.get("url") or "")
        category_hint = _infer_category(title)
        entity_hint = _infer_entity(f"notion_{title}_{page_id}")
        chunks: list[str] = []
        properties = page.get("properties")
        if isinstance(properties, dict):
            chunks.extend(_flatten_notion_properties(properties))

        block_chunks, child_pages = self._extract_page_content(page_id)
        chunks.extend(block_chunks)

        records: list[dict[str, Any]] = []
        limited_chunks = chunks[: self.max_records_per_page]
        for idx, chunk in enumerate(limited_chunks, start=1):
            content = _normalize_whitespace(chunk)
            if not content:
                continue
            if len(content) > self.max_chars_per_record:
                content = f"{content[:self.max_chars_per_record]}...(truncated)"
            fingerprint = sha256(content.encode("utf-8")).hexdigest()
            if fingerprint in seen_fingerprints:
                continue
            seen_fingerprints.add(fingerprint)
            source_id = f"notion:{page_id}:{idx}:{fingerprint[:10]}"
            records.append(
                {
                    "source_id": source_id,
                    "content": content,
                    "entity_key": entity_hint,
                    "category": category_hint,
                    "metadata": {
                        "legacy_import": True,
                        "source_connector": "notion",
                        "notion_page_id": page_id,
                        "notion_title": title,
                        "notion_url": page_url,
                        "last_edited_time": page.get("last_edited_time"),
                        "chunk_index": idx,
                        "fingerprint": fingerprint[:16],
                    },
                }
            )
        return records, child_pages

    def _extract_page_content(self, page_id: str) -> tuple[list[str], list[str]]:
        visited_blocks: set[str] = set()
        return self._extract_block_children(page_id, visited_blocks=visited_blocks, depth=0)

    def _extract_block_children(
        self,
        block_id: str,
        *,
        visited_blocks: set[str],
        depth: int,
    ) -> tuple[list[str], list[str]]:
        if depth > 8:
            self._warnings.append(f"notion:block:{block_id}: depth_limit_reached")
            return [], []
        if block_id in visited_blocks:
            return [], []
        visited_blocks.add(block_id)

        try:
            blocks = self.client.get_block_children(block_id)
        except Exception as exc:
            self._warnings.append(f"notion:block:{block_id}: fetch_error:{type(exc).__name__}")
            return [], []

        chunks: list[str] = []
        child_pages: list[str] = []
        for block in blocks:
            block_id_raw = str(block.get("id", ""))
            block_type = str(block.get("type", ""))
            block_text = _notion_block_text(block)
            if block_text:
                chunks.append(block_text)
            if block_type == "child_page" and block_id_raw:
                child_pages.append(_normalize_notion_id(block_id_raw))
            if bool(block.get("has_children")) and block_id_raw:
                nested_chunks, nested_pages = self._extract_block_children(
                    _normalize_notion_id(block_id_raw),
                    visited_blocks=visited_blocks,
                    depth=depth + 1,
                )
                chunks.extend(nested_chunks)
                child_pages.extend(nested_pages)
        return chunks, child_pages


def _split_paragraphs(text: str) -> list[str]:
    parts = re.split(r"\n\s*\n", text)
    return [_normalize_whitespace(item) for item in parts if _normalize_whitespace(item)]


def _flatten_payload(payload: Any, *, prefix: str = "") -> list[str]:
    if payload is None:
        return []
    if isinstance(payload, str):
        return [payload.strip()] if payload.strip() else []
    if isinstance(payload, (int, float, bool)):
        return [f"{prefix}{payload}".strip()]
    if isinstance(payload, list):
        out: list[str] = []
        for idx, item in enumerate(payload):
            child_prefix = f"{prefix}[{idx}] " if prefix else f"[{idx}] "
            out.extend(_flatten_payload(item, prefix=child_prefix))
        return out
    if isinstance(payload, dict):
        out: list[str] = []
        for key, value in payload.items():
            child_prefix = f"{prefix}{key}: " if prefix else f"{key}: "
            out.extend(_flatten_payload(value, prefix=child_prefix))
        return out
    text = str(payload).strip()
    return [f"{prefix}{text}".strip()] if text else []


def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _normalize_notion_id(value: str) -> str:
    candidate = value.strip()
    if not candidate:
        return ""
    return candidate.replace("-", "")


def _notion_page_title(page: dict[str, Any]) -> str:
    properties = page.get("properties")
    if isinstance(properties, dict):
        for prop in properties.values():
            if not isinstance(prop, dict):
                continue
            if prop.get("type") != "title":
                continue
            title_fragments = prop.get("title")
            if isinstance(title_fragments, list):
                text = _notion_rich_text_to_plain(title_fragments)
                if text:
                    return text
    title_fallback = page.get("title")
    if isinstance(title_fallback, list):
        return _notion_rich_text_to_plain(title_fallback)
    return ""


def _flatten_notion_properties(properties: dict[str, Any]) -> list[str]:
    out: list[str] = []
    for key, prop in properties.items():
        if not isinstance(prop, dict):
            continue
        prop_type = str(prop.get("type") or "")
        value = _notion_property_value(prop_type, prop.get(prop_type))
        value = _normalize_whitespace(value)
        if not value:
            continue
        out.append(f"property {key}: {value}")
    return out


def _notion_property_value(prop_type: str, value: Any) -> str:
    if value is None:
        return ""
    if prop_type in {"title", "rich_text"} and isinstance(value, list):
        return _notion_rich_text_to_plain(value)
    if prop_type == "select" and isinstance(value, dict):
        return str(value.get("name") or "")
    if prop_type == "multi_select" and isinstance(value, list):
        return ", ".join(str(item.get("name") or "") for item in value if isinstance(item, dict) and item.get("name"))
    if prop_type in {"date", "created_time", "last_edited_time"}:
        if isinstance(value, dict):
            start = value.get("start")
            end = value.get("end")
            if start and end:
                return f"{start} -> {end}"
            return str(start or "")
        return str(value)
    if prop_type == "number":
        return str(value)
    if prop_type == "checkbox":
        return "true" if bool(value) else "false"
    if prop_type == "url":
        return str(value)
    if prop_type == "email":
        return str(value)
    if prop_type == "phone_number":
        return str(value)
    if prop_type in {"people", "relation"} and isinstance(value, list):
        ids: list[str] = []
        for item in value:
            if isinstance(item, dict):
                item_id = item.get("id")
                if item_id:
                    ids.append(str(item_id))
        return ", ".join(ids)
    if prop_type in {"formula", "rollup"} and isinstance(value, dict):
        value_type = str(value.get("type") or "")
        return _notion_property_value(value_type, value.get(value_type))
    if isinstance(value, dict):
        if "plain_text" in value:
            return str(value.get("plain_text") or "")
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _notion_rich_text_to_plain(items: list[Any]) -> str:
    out: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        plain = item.get("plain_text")
        if isinstance(plain, str) and plain.strip():
            out.append(plain)
    return _normalize_whitespace("".join(out))


def _notion_block_text(block: dict[str, Any]) -> str:
    block_type = str(block.get("type") or "")
    if not block_type:
        return ""
    payload = block.get(block_type)
    if not isinstance(payload, dict):
        payload = {}

    if block_type in {
        "paragraph",
        "heading_1",
        "heading_2",
        "heading_3",
        "bulleted_list_item",
        "numbered_list_item",
        "to_do",
        "toggle",
        "quote",
        "callout",
        "code",
    }:
        text = _notion_rich_text_to_plain(payload.get("rich_text") or [])
        if not text:
            return ""
        if block_type == "to_do":
            checked = bool(payload.get("checked"))
            return f"todo [{'x' if checked else ' '}] {text}"
        if block_type == "code":
            language = payload.get("language")
            if language:
                return f"code ({language}): {text}"
        return text

    if block_type == "child_page":
        title = str(payload.get("title") or "").strip()
        return f"child page: {title}" if title else ""

    if block_type == "table_row":
        cells = payload.get("cells")
        if not isinstance(cells, list):
            return ""
        rendered_cells = []
        for cell in cells:
            if isinstance(cell, list):
                cell_text = _notion_rich_text_to_plain(cell)
                if cell_text:
                    rendered_cells.append(cell_text)
        return " | ".join(rendered_cells)

    if block_type in {"bookmark", "embed", "link_preview"}:
        url = payload.get("url")
        return str(url or "")
    if block_type in {"image", "video", "file", "pdf"}:
        file_meta = payload.get("file")
        external_meta = payload.get("external")
        if isinstance(file_meta, dict) and file_meta.get("url"):
            return str(file_meta.get("url"))
        if isinstance(external_meta, dict) and external_meta.get("url"):
            return str(external_meta.get("url"))
        return ""
    if block_type == "equation":
        expr = payload.get("expression")
        return f"equation: {expr}" if expr else ""

    text = _notion_rich_text_to_plain(payload.get("rich_text") or [])
    if text:
        return text
    plain = payload.get("plain_text")
    if isinstance(plain, str):
        return _normalize_whitespace(plain)
    return ""


def _infer_category(path_text: str) -> str:
    normalized = path_text.lower()
    mapping = {
        "support": "support_policy",
        "ticket": "support_policy",
        "warehouse": "warehouse_policy",
        "logistics": "operations",
        "driver": "operations",
        "customer": "customer_preferences",
        "access": "access_policy",
        "gate": "access_policy",
    }
    for key, category in mapping.items():
        if key in normalized:
            return category
    return "legacy_import"


def _infer_entity(path_text: str) -> str:
    normalized = re.sub(r"[^\w]+", "_", path_text.lower()).strip("_")
    return normalized[:120] or "legacy_entity"


def _slugify_key(value: str, *, max_len: int = 120) -> str:
    candidate = value.strip().lower()
    if not candidate:
        return ""
    candidate = candidate.replace("/", " ")
    candidate = re.sub(r"[^\w\s-]+", "_", candidate)
    candidate = re.sub(r"[\s-]+", "_", candidate)
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate:
        return ""
    return candidate[:max_len]


def _humanize_token(value: str) -> str:
    token = _normalize_whitespace(re.sub(r"[_-]+", " ", value)).strip()
    if not token:
        return ""
    return token[:1].upper() + token[1:]


def _normalize_seed_category(value: str) -> str:
    normalized = _slugify_key(value, max_len=80)
    if not normalized:
        return ""
    alias_map = {
        "ops": "operations",
        "operation": "operations",
        "operations_policy": "operations",
        "operatsii": "operations",
        "operaciya": "operations",
        "операции": "operations",
        "операция": "operations",
        "customer": "customer_preferences",
        "customers": "customer_preferences",
        "client": "customer_preferences",
        "clients": "customer_preferences",
        "клиент": "customer_preferences",
        "клиенты": "customer_preferences",
        "клиентские_предпочтения": "customer_preferences",
        "support": "support_policy",
        "поддержка": "support_policy",
        "warehouse": "warehouse_policy",
        "склад": "warehouse_policy",
        "access": "access_policy",
        "access_policy_rules": "access_policy",
        "доступ": "access_policy",
        "ворота": "access_policy",
    }
    return alias_map.get(normalized, normalized)


def _normalize_section_overrides(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        return {}
    out: dict[str, Any] = {}
    for raw_key, raw_override in value.items():
        key = _normalize_seed_category(str(raw_key))
        if not key:
            key = _slugify_key(str(raw_key), max_len=80)
        if not key:
            continue
        out[key] = raw_override
    return out


def _infer_seed_entity(*, metadata: dict[str, Any], source_id: str) -> str:
    candidates: list[str] = []
    explicit = metadata.get("entity_key")
    if explicit:
        candidates.append(str(explicit))

    notion_title = metadata.get("notion_title")
    if notion_title:
        candidates.append(str(notion_title))

    file_path = str(metadata.get("file_path") or "").strip()
    if file_path:
        try:
            file_stem = Path(file_path).stem
            if file_stem:
                candidates.append(file_stem)
        except Exception:
            candidates.append(file_path)

    notion_page_id = metadata.get("notion_page_id")
    if notion_page_id:
        candidates.append(f"notion_{notion_page_id}")

    if source_id:
        candidates.append(source_id.split(":")[0])
        candidates.append(source_id)

    for candidate in candidates:
        slug = _slugify_key(str(candidate), max_len=120)
        if slug:
            return slug
    return "legacy_entity"


def _infer_seed_space(*, source_type: str, category: str, metadata: dict[str, Any]) -> str:
    explicit = _slugify_key(str(metadata.get("space_key") or ""))
    if explicit:
        return explicit

    normalized_source = _slugify_key(source_type, max_len=50)
    source_connector = _slugify_key(str(metadata.get("source_connector") or ""), max_len=50)
    if source_connector == "notion" or normalized_source.startswith("notion"):
        return "notion"
    if normalized_source in {"local_dir", "local", "file"}:
        return "files"

    category_space_map = {
        "support_policy": "support",
        "warehouse_policy": "warehouse",
        "customer_preferences": "customers",
        "access_policy": "access",
        "operations": "operations",
        "legacy_import": "legacy",
    }
    return category_space_map.get(category, "legacy")


def _resolve_seed_section(*, category: str, overrides: dict[str, Any]) -> tuple[str, str]:
    default_map: dict[str, tuple[str, str]] = {
        "access_policy": ("policies", "Access Policy"),
        "warehouse_policy": ("operations", "Warehouse Operations"),
        "support_policy": ("support", "Support Rules"),
        "customer_preferences": ("preferences", "Customer Preferences"),
        "operations": ("operations", "Operations"),
        "legacy_import": ("notes", "Imported Notes"),
    }
    base_key, base_heading = default_map.get(category, ("notes", _humanize_token(category) or "Imported Notes"))
    override = overrides.get(category)
    if isinstance(override, str):
        heading = _normalize_whitespace(override) or base_heading
        section_key = _slugify_key(override, max_len=60) or base_key
        return section_key, heading
    if isinstance(override, dict):
        key_candidate = str(override.get("section_key") or override.get("key") or "")
        heading_candidate = str(override.get("section_heading") or override.get("heading") or "")
        section_key = _slugify_key(key_candidate, max_len=60) or base_key
        heading = _normalize_whitespace(heading_candidate) or base_heading
        return section_key, heading
    return base_key, base_heading


def _build_seed_page_token(
    *,
    group_mode: str,
    entity_key: str,
    category: str,
    metadata: dict[str, Any],
) -> str:
    if group_mode == "entity":
        return _slugify_key(entity_key, max_len=140) or "legacy_entity"
    if group_mode == "category":
        return _slugify_key(category, max_len=140) or "legacy_category"
    if group_mode == "category_entity":
        return _slugify_key(f"{category}_{entity_key}", max_len=140) or "legacy_page"

    # Defensive fallback for unknown future modes.
    fallback = _slugify_key(str(metadata.get("notion_title") or ""), max_len=140)
    return fallback or _slugify_key(entity_key, max_len=140) or "legacy_page"


def _join_slug(page_prefix: str, space_key: str, page_token: str) -> str:
    parts = [
        _slugify_key(page_prefix, max_len=80),
        _slugify_key(space_key, max_len=80),
        _slugify_key(page_token, max_len=160),
    ]
    normalized = [part for part in parts if part]
    if not normalized:
        return "legacy"
    return "/".join(normalized)


def _build_seed_page_title(
    *,
    entity_key: str,
    category: str,
    metadata: dict[str, Any],
    section_heading: str,
) -> str:
    notion_title = _normalize_whitespace(str(metadata.get("notion_title") or ""))
    if notion_title:
        base_title = notion_title
    else:
        file_path = str(metadata.get("file_path") or "").strip()
        file_stem = Path(file_path).stem if file_path else ""
        base_title = _humanize_token(file_stem) or _humanize_token(entity_key) or "Legacy Knowledge"

    category_label = _humanize_token(category) or "Knowledge"
    heading_label = _normalize_whitespace(section_heading)
    if category_label.lower() in base_title.lower():
        title = base_title
    else:
        title = f"{base_title} ({category_label})"
    if heading_label and heading_label.lower() not in title.lower():
        title = f"{title} - {heading_label}"
    return title[:180]
